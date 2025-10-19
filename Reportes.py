from flask import Blueprint, jsonify, request, send_file
from db import mysql
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

reportes_bp = Blueprint('reportes_bp', __name__)

@reportes_bp.route('/api/reporte_alertas', methods=['GET'])
def generar_reporte_alertas():
    """
    Genera un reporte PDF con todas las alertas almacenadas en la base de datos.
    """
    try:
        # 1️⃣ Consultar datos de la tabla Alertas
        cursor = mysql.connection.cursor()
        cursor.execute("""
            SELECT Cuencas, Rios, Niveles, Condiciones, Pronosticos, Periodos
            FROM Alertas
        """)
        alertas = cursor.fetchall()
        cursor.close()

        if not alertas:
            return jsonify({'mensaje': 'No hay alertas registradas para generar el reporte.'}), 404

        # 2️⃣ Crear archivo PDF temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        doc = SimpleDocTemplate(
            temp_file.name,
            pagesize=letter,
            leftMargin=0.8*inch,
            rightMargin=0.8*inch,
            topMargin=0.8*inch,
            bottomMargin=0.8*inch
        )

        elements = []
        styles = getSampleStyleSheet()

        # 🔹 Estilo personalizado centrado
        estilo_titulo = ParagraphStyle(
            name="TituloCentrado",
            parent=styles["Title"],
            alignment=1,  # 0=izq, 1=centro, 2=der
            textColor=colors.black
        )

        estilo_pie = ParagraphStyle(
            name="PieCentrado",
            parent=styles["Normal"],
            alignment=1,
            textColor=colors.gray
        )

        # 3️⃣ Título del reporte
        titulo = Paragraph("📘 Reporte de Alertas - HidroAlert", estilo_titulo)
        elements.append(titulo)
        elements.append(Spacer(1, 0.3 * inch))

        # 4️⃣ Crear tabla con los datos
        encabezados = ['Cuenca', 'Río', 'Nivel', 'Condición', 'Pronóstico', 'Periodo']
        data = [encabezados] + list(alertas)

        tabla = Table(data, colWidths=[1.2*inch]*6)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#007BFF")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ]))

        elements.append(tabla)
        elements.append(Spacer(1, 0.3 * inch))

        # 5️⃣ Pie de página
        pie = Paragraph("Generado automáticamente por el sistema HidroAlert © 2025", estilo_pie)
        elements.append(pie)

        # 6️⃣ Generar PDF
        doc.build(elements)

        # 7️⃣ Enviar archivo al cliente
        return send_file(temp_file.name, as_attachment=True, download_name='reporte_alertas.pdf')

    except Exception as ex:
        return jsonify({'mensaje': 'Error al generar el reporte ❌', 'error': str(ex)}), 500
    finally:
        try:
            cursor.close()
        except:
            pass



# ✅ NUEVA API: GENERAR REPORTE EN EXCEL
@reportes_bp.route('/api/reporte_alertas_excel', methods=['GET'])
def generar_reporte_alertas_excel():
    """
    Genera un reporte Excel (.xlsx) con todas las alertas almacenadas en la base de datos.
    """
    try:
        # 1️⃣ Consultar datos de la tabla Alertas
        cursor = mysql.connection.cursor()
        cursor.execute("""
            SELECT Cuencas, Rios, Niveles, Condiciones, Pronosticos, Periodos
            FROM Alertas
        """)
        alertas = cursor.fetchall()
        cursor.close()

        if not alertas:
            return jsonify({'mensaje': 'No hay alertas registradas para generar el reporte.'}), 404

        # 2️⃣ Crear libro de Excel temporal
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Alertas"

        # 3️⃣ Encabezados
        encabezados = ['Cuenca', 'Río', 'Nivel', 'Condición', 'Pronóstico', 'Periodo']
        ws.append(encabezados)

        # Estilos de encabezado
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin', color='999999'),
                        right=Side(style='thin', color='999999'),
                        top=Side(style='thin', color='999999'),
                        bottom=Side(style='thin', color='999999'))

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(encabezados)):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border

        # 4️⃣ Agregar datos
        for fila in alertas:
            ws.append(fila)

        # Aplicar formato a celdas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
            for cell in row:
                cell.alignment = center_align
                cell.border = border

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 5️⃣ Pie de página
        ws.append([])
        ws.append(["Generado automáticamente por el sistema HidroAlert © 2025"])
        ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
        ws.cell(row=ws.max_row, column=1).alignment = center_align
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="888888")

        # 6️⃣ Guardar archivo
        wb.save(temp_file.name)

        # 7️⃣ Enviar archivo al cliente
        return send_file(temp_file.name, as_attachment=True, download_name='reporte_alertas.xlsx')

    except Exception as ex:
        return jsonify({'mensaje': 'Error al generar el reporte Excel ❌', 'error': str(ex)}), 500
    finally:
        try:
            cursor.close()
        except:
            pass
